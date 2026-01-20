"""
Radio Classics Schedule Fetcher
Scrapes gregbellmedia.com for the weekly SiriusXM Channel 148 schedule,
parses the Excel file, and outputs JSON for the website.
"""

import json
import logging
import re
import sys
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constants
SCHEDULE_SOURCE_URL = "https://gregbellmedia.com/"
OUTPUT_PATH = Path(__file__).parent.parent / "docs" / "schedule.json"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"


def fetch_page(url: str) -> Optional[str]:
    """Fetch HTML content from a URL."""
    try:
        response = requests.get(
            url,
            headers={"User-Agent": USER_AGENT},
            timeout=30
        )
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        logger.error(f"Failed to fetch {url}: {e}")
        return None


def parse_date_from_filename(filename: str) -> Optional[tuple[datetime, datetime]]:
    """Parse date range from Excel filename.

    Handles formats like:
    - RC_Jan12th2026-Jan18th2026-Excel-Version.xlsx
    - RC_Jan5th2026-Jan11th2026-Excel-Version.xlsx

    Returns: (start_date, end_date) as datetime objects, or None if parsing fails.
    """
    MONTHS = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }

    # Pattern: MonthDay(st/nd/rd/th)Year-MonthDay(st/nd/rd/th)Year
    pattern = re.compile(
        r'([A-Za-z]{3})(\d{1,2})(?:st|nd|rd|th)?(\d{4})-'
        r'([A-Za-z]{3})(\d{1,2})(?:st|nd|rd|th)?(\d{4})',
        re.IGNORECASE
    )

    match = pattern.search(filename)
    if not match:
        return None

    try:
        start_month = MONTHS.get(match.group(1).lower())
        start_day = int(match.group(2))
        start_year = int(match.group(3))

        end_month = MONTHS.get(match.group(4).lower())
        end_day = int(match.group(5))
        end_year = int(match.group(6))

        if not start_month or not end_month:
            return None

        start_date = datetime(start_year, start_month, start_day)
        end_date = datetime(end_year, end_month, end_day)

        return start_date, end_date
    except (ValueError, TypeError):
        return None


def find_excel_url(html: str) -> Optional[str]:
    """Extract the Excel schedule URL for the current week from the page HTML.

    Finds all Excel files, parses their date ranges, and selects the one
    that contains today's date. Falls back to the most recent if current
    week is not found.
    """
    soup = BeautifulSoup(html, 'html.parser')
    today = datetime.now()

    # Pattern for Radio Classics Excel files
    excel_pattern = re.compile(r'RC_.*Excel.*\.xlsx', re.IGNORECASE)

    # Collect all Excel URLs with their date ranges
    excel_files = []

    for link in soup.find_all('a', href=True):
        href = link['href']
        if excel_pattern.search(href):
            # Handle protocol-relative URLs
            if href.startswith('//'):
                href = 'https:' + href

            # Try to parse dates from filename
            date_range = parse_date_from_filename(href)
            if date_range:
                start_date, end_date = date_range
                excel_files.append({
                    'url': href,
                    'start': start_date,
                    'end': end_date
                })
                logger.debug(f"Found Excel: {href} ({start_date.date()} to {end_date.date()})")
            else:
                # Include files without parseable dates as fallback
                excel_files.append({
                    'url': href,
                    'start': None,
                    'end': None
                })
                logger.debug(f"Found Excel (no date): {href}")

    if not excel_files:
        # Fallback: look for any .xlsx files
        for link in soup.find_all('a', href=True):
            href = link['href']
            if href.endswith('.xlsx'):
                if href.startswith('//'):
                    href = 'https:' + href
                logger.info(f"Found Excel URL (fallback): {href}")
                return href

        logger.error("No Excel file URL found on the page")
        return None

    logger.info(f"Found {len(excel_files)} Excel files on the page")

    # First pass: find file containing today's date
    for excel in excel_files:
        if excel['start'] and excel['end']:
            if excel['start'].date() <= today.date() <= excel['end'].date():
                logger.info(f"Selected Excel for current week: {excel['url']}")
                logger.info(f"  Date range: {excel['start'].date()} to {excel['end'].date()}")
                return excel['url']

    # Second pass: find the most recent file (by end date)
    dated_files = [f for f in excel_files if f['start'] is not None]
    if dated_files:
        # Sort by end date descending
        dated_files.sort(key=lambda x: x['end'], reverse=True)
        best = dated_files[0]
        logger.warning(f"Current week not found, using most recent: {best['url']}")
        logger.warning(f"  Date range: {best['start'].date()} to {best['end'].date()}")
        return best['url']

    # Last resort: return first Excel file found
    logger.warning(f"No dated files found, using first Excel: {excel_files[0]['url']}")
    return excel_files[0]['url']


def download_excel(url: str) -> Optional[BytesIO]:
    """Download Excel file and return as BytesIO."""
    try:
        response = requests.get(
            url,
            headers={"User-Agent": USER_AGENT},
            timeout=60
        )
        response.raise_for_status()
        logger.info(f"Downloaded Excel file ({len(response.content)} bytes)")
        return BytesIO(response.content)
    except requests.RequestException as e:
        logger.error(f"Failed to download Excel file: {e}")
        return None


def format_time_et(hour: int, minute: int) -> str:
    """Format hour (0-23) and minute as '12:00 AM' style string."""
    if hour == 0:
        display_hour = 12
        period = "AM"
    elif hour < 12:
        display_hour = hour
        period = "AM"
    elif hour == 12:
        display_hour = 12
        period = "PM"
    else:
        display_hour = hour - 12
        period = "PM"

    return f"{display_hour}:{minute:02d} {period}"


def parse_time_value_to_hour(time_val: str) -> Optional[int]:
    """Parse a time value from the ET column into a 24-hour value.

    Handles formats like: '12mid', '2am', '4am', '12noon', '2pm', '10pm'
    Returns the hour (0-23) or None if not recognized.
    """
    if not time_val:
        return None

    time_str = str(time_val).lower().strip()

    # Handle special cases
    if 'mid' in time_str or time_str == '12am':
        return 0
    if 'noon' in time_str or time_str == '12pm':
        return 12

    # Pattern: number followed by am/pm
    match = re.match(r'(\d{1,2})\s*(am|pm)', time_str)
    if match:
        hour = int(match.group(1))
        period = match.group(2)
        if period == 'am':
            return hour if hour != 12 else 0
        else:  # pm
            return hour if hour == 12 else hour + 12

    return None


# Words that indicate a cell is a continuation of the previous cell
CONTINUATION_WORDS = ['from ', 'with ', 'starring ', 'featuring ', 'hosted by ', 'and ']

# Known 60-minute shows (partial matches)
HOUR_SHOWS = ['lux radio', 'screen director', 'theatre guild', 'screen guild']


def get_block_for_row(row: int, time_blocks: list[tuple[int, int]],
                      rows_per_block: int) -> int:
    """Get the block index for a given row number."""
    for i, (block_start, _) in enumerate(time_blocks):
        block_end = block_start + rows_per_block
        if block_start <= row < block_end:
            return i
    return -1


def is_continuation(text: str, prev_text: str, row: int, prev_row: int,
                    time_blocks: list[tuple[int, int]], rows_per_block: int) -> bool:
    """Determine if text is a continuation of the previous cell.

    Returns True if the text should be joined with the previous cell.
    """
    if not text or not prev_text:
        return False

    # Never join across time block boundaries
    if get_block_for_row(row, time_blocks, rows_per_block) != \
       get_block_for_row(prev_row, time_blocks, rows_per_block):
        return False

    text_lower = text.lower().strip()
    has_date = bool(re.search(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', text))

    # Starts with continuation words = definitely continuation
    for word in CONTINUATION_WORDS:
        if text_lower.startswith(word):
            return True

    # Lowercase start WITHOUT a date = continuation
    if text and text[0].islower() and not has_date:
        return True

    # Two consecutive cells without dates (excluding articles) = continuation
    if not has_date and not any(text_lower.startswith(x) for x in ['the ', 'a ', 'an ']):
        prev_has_date = bool(re.search(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', prev_text))
        if not prev_has_date:
            # Check if prev_text looks like a show name that continues
            # Avoid joining unrelated shows - only join if no year pattern in prev
            return True

    return False


def estimate_show_duration(show_name: str) -> int:
    """Estimate show duration in minutes based on name patterns.

    Old-time radio shows were typically 15, 30, or 60 minutes.
    Returns duration in minutes.
    """
    show_lower = show_name.lower()

    # Theme headers don't consume time (they're descriptive)
    if show_lower.startswith('[theme]'):
        return 0

    # Explicit duration markers in the data
    if '(1 hr)' in show_lower or '(60 min)' in show_lower or '(1 hour)' in show_lower:
        return 60
    if '(1/2 hr)' in show_lower or '(30 min)' in show_lower:
        return 30
    if '(15 min)' in show_lower:
        return 15

    # "Two From X" or "Two X Episodes" = 2 episodes x 30 min
    if show_lower.startswith('two from ') or 'two episodes' in show_lower:
        return 60

    # "Two 1/2 Hour" pattern = 2 x 30 min
    if 'two 1/2 hour' in show_lower:
        return 60

    # Known 60-minute shows
    if any(show in show_lower for show in HOUR_SHOWS):
        return 60

    # Default to 30 minutes (most common for Golden Age radio)
    return 30


def is_theme_header(show_name: str, is_bold: bool) -> bool:
    """Check if this is a theme header (birthday, marathon, etc.).

    Theme headers are bold cells without dates that indicate themed programming.
    """
    has_date = bool(re.search(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', show_name))
    if is_bold and not has_date:
        lower = show_name.lower()
        theme_keywords = ['birthday', 'marathon', 'when radio was', 'tribute',
                         'anniversary', 'salute', 'celebration', 'memorial']
        if any(kw in lower for kw in theme_keywords):
            return True
    return False


def find_header_row(ws, max_rows: int = 20) -> Optional[tuple[int, dict]]:
    """Find the header row containing day names and return column mapping.

    Returns: (row_number, {col_index: day_name}) or None if not found.
    """
    day_names = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']

    for row_idx in range(1, max_rows + 1):
        day_columns = {}

        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                cell_str = str(cell_val).lower().strip()
                for day in day_names:
                    if day in cell_str:
                        # Capitalize properly for output
                        day_columns[col_idx] = day.capitalize()
                        break

        # Need at least 5 days to consider this the header row
        if len(day_columns) >= 5:
            logger.info(f"Found header row at row {row_idx} with {len(day_columns)} days")
            return row_idx, day_columns

    return None


def find_et_column(ws, header_row: int) -> Optional[int]:
    """Find the Eastern Time (ET) column in the header row."""
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col_idx).value
        if cell_val:
            cell_str = str(cell_val).upper().strip()
            if cell_str == 'ET' or 'EASTERN' in cell_str:
                logger.info(f"Found ET column at column {col_idx}")
                return col_idx
    return None


def detect_time_blocks(ws, et_column: int, start_row: int, max_rows: int = 100) -> list[tuple[int, int]]:
    """Detect time blocks by reading the ET column.

    Returns list of (row_number, hour_24h) for each time block start.
    """
    time_blocks = []

    for row_idx in range(start_row, start_row + max_rows):
        cell_val = ws.cell(row=row_idx, column=et_column).value
        if cell_val:
            hour = parse_time_value_to_hour(cell_val)
            if hour is not None:
                time_blocks.append((row_idx, hour))
                logger.debug(f"Found time block at row {row_idx}: {cell_val} -> {hour}:00")

    return time_blocks


def parse_excel_schedule(excel_data: BytesIO) -> dict:
    """Parse the Excel schedule into structured JSON format.

    Dynamically detects:
    - Header row (by finding day names: Monday, Tuesday, etc.)
    - Day columns (by their position in the header)
    - ET column (by finding "ET" in header)
    - Time blocks (by reading time values from ET column)
    - Rows per block (by counting rows between time markers)

    Also handles:
    - Multi-line show descriptions (joins continuations)
    - Themed programming (bold cells without dates)
    - Variable show durations (15, 30, or 60 minutes)
    """
    # Load workbook twice: once for values, once for formatting (bold detection)
    wb_values = load_workbook(excel_data, data_only=True)
    ws_values = wb_values.active

    excel_data.seek(0)  # Reset stream position
    wb_format = load_workbook(excel_data, data_only=False)
    ws_format = wb_format.active

    # Step 1: Find the header row with day names
    header_result = find_header_row(ws_values)
    if not header_result:
        logger.error("Could not find header row with day names")
        raise ValueError("Could not find header row with day names (MONDAY, TUESDAY, etc.)")

    header_row, day_columns = header_result
    logger.info(f"Day columns detected: {day_columns}")

    # Step 2: Find the ET (Eastern Time) column
    et_column = find_et_column(ws_values, header_row)
    if not et_column:
        logger.warning("Could not find ET column, will try to infer time from block positions")

    # Step 3: Detect time blocks from the ET column
    data_start_row = header_row + 1
    time_blocks = []

    if et_column:
        time_blocks = detect_time_blocks(ws_values, et_column, data_start_row)
        logger.info(f"Detected {len(time_blocks)} time blocks from ET column")

    # Step 4: If we found time blocks, calculate rows per block
    if len(time_blocks) >= 2:
        rows_per_block = time_blocks[1][0] - time_blocks[0][0]
        logger.info(f"Rows per block: {rows_per_block}")

        for i in range(1, len(time_blocks)):
            gap = time_blocks[i][0] - time_blocks[i-1][0]
            if gap != rows_per_block:
                logger.warning(f"Inconsistent block spacing: expected {rows_per_block}, got {gap} at block {i}")
    else:
        logger.warning("Could not detect time blocks, using default structure (5 rows)")
        rows_per_block = 5
        time_blocks = [(data_start_row + i * rows_per_block, i * 2) for i in range(12)]

    # Step 5: Initialize day schedules
    day_schedules = {day: [] for day in day_columns.values()}

    # Step 6: Process each day column separately to handle continuations
    for col_idx, day_name in day_columns.items():
        # Collect all cells for this day with their row numbers and values
        day_cells = []

        for block_idx, (block_start_row, base_hour) in enumerate(time_blocks):
            for slot_idx in range(rows_per_block):
                row_num = block_start_row + slot_idx

                cell_value = ws_values.cell(row=row_num, column=col_idx).value
                cell_format = ws_format.cell(row=row_num, column=col_idx)

                # Check if cell is bold
                cell_bold = False
                if cell_format.font and cell_format.font.bold:
                    cell_bold = True

                if cell_value and str(cell_value).strip():
                    val = str(cell_value).strip()
                    if val.lower() not in ['none', 'n/a', '-', '']:
                        day_cells.append({
                            'row': row_num,
                            'value': val,
                            'bold': cell_bold,
                            'block_idx': block_idx,
                            'base_hour': base_hour
                        })

        # Step 7: Join continuations for this day
        joined_shows = []
        i = 0

        while i < len(day_cells):
            current = day_cells[i]
            show_parts = [current['value']]
            start_row = current['row']
            start_block_idx = current['block_idx']
            start_base_hour = current['base_hour']
            is_bold = current['bold']

            # Look ahead for continuations
            j = i + 1
            while j < len(day_cells):
                next_cell = day_cells[j]
                prev_text = show_parts[-1]

                if is_continuation(next_cell['value'], prev_text, next_cell['row'],
                                   day_cells[j-1]['row'], time_blocks, rows_per_block):
                    show_parts.append(next_cell['value'])
                    j += 1
                else:
                    break

            # Build the joined show name
            show_name = ' '.join(show_parts)

            # Check if this is a theme header
            if is_theme_header(show_name, is_bold):
                show_name = f"[THEME] {show_name}"

            joined_shows.append({
                'show': show_name,
                'block_idx': start_block_idx,
                'base_hour': start_base_hour,
                'start_row': start_row
            })

            i = j

        # Step 8: Calculate times based on show durations within each block
        # Group shows by block
        blocks = {}
        for show in joined_shows:
            block_idx = show['block_idx']
            if block_idx not in blocks:
                blocks[block_idx] = []
            blocks[block_idx].append(show)

        # Process each block and assign times
        for block_idx, block_shows in blocks.items():
            if not block_shows:
                continue

            base_hour = block_shows[0]['base_hour']
            current_minutes = base_hour * 60

            for show in block_shows:
                hour = (current_minutes // 60) % 24
                minute = current_minutes % 60
                time_str = format_time_et(hour, minute)

                day_schedules[day_name].append({
                    "time": time_str,
                    "show": show['show'],
                    "episode": ""
                })

                # Add duration for next show's time
                duration = estimate_show_duration(show['show'])
                current_minutes += duration

        logger.debug(f"  {day_name}: {len(day_cells)} cells -> {len(joined_shows)} shows")

    # Step 9: Log results
    total_shows = sum(len(slots) for slots in day_schedules.values())
    logger.info(f"Parsed {total_shows} total shows across all days (after joining continuations)")
    for day_name, slots in day_schedules.items():
        logger.info(f"  {day_name}: {len(slots)} shows")

    # Extract date range from the workbook
    week_start, week_end = extract_date_range(ws_values)

    schedule_data = {
        "week_start": week_start,
        "week_end": week_end,
        "last_updated": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "schedule": []
    }

    # Build final schedule structure (Sunday first for display consistency)
    for day in ['Sunday', 'Monday', 'Tuesday', 'Wednesday',
                'Thursday', 'Friday', 'Saturday']:
        if day in day_schedules:
            schedule_data["schedule"].append({
                "day": day,
                "date": "",  # Will be filled by frontend based on week_start
                "slots": sorted(day_schedules[day],
                              key=lambda x: parse_time_for_sort(x["time"]))
            })

    return schedule_data


def parse_time_for_sort(time_str: str) -> int:
    """Convert time string to minutes for sorting."""
    try:
        time_str = time_str.upper().strip()

        # Handle AM/PM
        is_pm = 'PM' in time_str
        time_str = time_str.replace('AM', '').replace('PM', '').strip()

        parts = time_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1]) if len(parts) > 1 else 0

        # Convert to 24-hour for sorting
        if is_pm and hours != 12:
            hours += 12
        elif not is_pm and hours == 12:
            hours = 0

        return hours * 60 + minutes
    except (ValueError, IndexError):
        return 0


def extract_date_range(ws) -> tuple[str, str]:
    """Try to extract the date range from the worksheet.

    Handles formats like:
    - "Jan 19th - Jan 25th, 2026"
    - "January 19 - January 25, 2026"
    - "1/19/2026 - 1/25/2026"
    """
    # Month name mapping
    MONTHS = {
        'jan': 1, 'january': 1,
        'feb': 2, 'february': 2,
        'mar': 3, 'march': 3,
        'apr': 4, 'april': 4,
        'may': 5,
        'jun': 6, 'june': 6,
        'jul': 7, 'july': 7,
        'aug': 8, 'august': 8,
        'sep': 9, 'sept': 9, 'september': 9,
        'oct': 10, 'october': 10,
        'nov': 11, 'november': 11,
        'dec': 12, 'december': 12
    }

    # Pattern for "Jan 19th - Jan 25th, 2026" or "January 19 - January 25, 2026"
    text_date_pattern = re.compile(
        r'([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?\s*[-–]\s*'
        r'([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?,?\s*(\d{4})',
        re.IGNORECASE
    )

    # Pattern for numeric dates: "1/19/2026 - 1/25/2026"
    numeric_date_pattern = re.compile(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})')

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value:
                val = str(cell.value)

                # Try text date pattern first
                text_match = text_date_pattern.search(val)
                if text_match:
                    start_month_str = text_match.group(1).lower()
                    start_day = int(text_match.group(2))
                    end_month_str = text_match.group(3).lower()
                    end_day = int(text_match.group(4))
                    year = int(text_match.group(5))

                    start_month = MONTHS.get(start_month_str, 1)
                    end_month = MONTHS.get(end_month_str, 1)

                    start_date = f"{year}-{start_month:02d}-{start_day:02d}"
                    end_date = f"{year}-{end_month:02d}-{end_day:02d}"

                    logger.info(f"Extracted date range: {start_date} to {end_date}")
                    return start_date, end_date

                # Try numeric date pattern
                matches = numeric_date_pattern.findall(val)
                if len(matches) >= 2:
                    start = matches[0]
                    end = matches[1]

                    year = int(start[2])
                    if year < 100:
                        year += 2000

                    start_date = f"{year}-{int(start[0]):02d}-{int(start[1]):02d}"

                    year = int(end[2])
                    if year < 100:
                        year += 2000
                    end_date = f"{year}-{int(end[0]):02d}-{int(end[1]):02d}"

                    logger.info(f"Extracted date range: {start_date} to {end_date}")
                    return start_date, end_date

    # Default to current week (Monday-based since Excel uses Monday start)
    logger.warning("Could not extract date range from Excel, using current week")
    today = datetime.now()
    # Find most recent Monday
    days_since_monday = today.weekday()
    week_start = today - timedelta(days=days_since_monday)
    week_end = week_start + timedelta(days=6)

    return week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d")


def create_default_schedule() -> dict:
    """Create a default/placeholder schedule structure."""
    today = datetime.now()
    days_since_sunday = (today.weekday() + 1) % 7
    week_start = today - timedelta(days=days_since_sunday)
    week_end = week_start + timedelta(days=6)

    return {
        "week_start": week_start.strftime("%Y-%m-%d"),
        "week_end": week_end.strftime("%Y-%m-%d"),
        "last_updated": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "schedule": [
            {"day": day, "date": "", "slots": []}
            for day in ['Sunday', 'Monday', 'Tuesday', 'Wednesday',
                       'Thursday', 'Friday', 'Saturday']
        ],
        "error": "Could not parse schedule. Please check back later."
    }


def save_schedule(schedule: dict, output_path: Path) -> bool:
    """Save schedule to JSON file."""
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(schedule, f, indent=2, ensure_ascii=False)
        logger.info(f"Schedule saved to {output_path}")
        return True
    except IOError as e:
        logger.error(f"Failed to save schedule: {e}")
        return False


def main() -> int:
    """Main entry point."""
    logger.info("Starting Radio Classics schedule fetch")

    # Step 1: Fetch the source page
    html = fetch_page(SCHEDULE_SOURCE_URL)
    if not html:
        logger.error("Failed to fetch source page")
        schedule = create_default_schedule()
        schedule["error"] = "Could not fetch schedule source page"
        save_schedule(schedule, OUTPUT_PATH)
        return 1

    # Step 2: Find the Excel URL
    excel_url = find_excel_url(html)
    if not excel_url:
        logger.error("Could not find Excel file URL")
        schedule = create_default_schedule()
        schedule["error"] = "Could not find schedule file"
        save_schedule(schedule, OUTPUT_PATH)
        return 1

    # Step 3: Download the Excel file
    excel_data = download_excel(excel_url)
    if not excel_data:
        logger.error("Failed to download Excel file")
        schedule = create_default_schedule()
        schedule["error"] = "Could not download schedule file"
        save_schedule(schedule, OUTPUT_PATH)
        return 1

    # Step 4: Parse the schedule
    try:
        schedule = parse_excel_schedule(excel_data)
        logger.info(f"Parsed schedule for week {schedule['week_start']} to {schedule['week_end']}")
    except Exception as e:
        logger.error(f"Failed to parse Excel file: {e}")
        schedule = create_default_schedule()
        schedule["error"] = f"Could not parse schedule: {str(e)}"
        save_schedule(schedule, OUTPUT_PATH)
        return 1

    # Step 5: Save the schedule
    if not save_schedule(schedule, OUTPUT_PATH):
        return 1

    logger.info("Schedule fetch completed successfully")
    return 0


if __name__ == "__main__":
    sys.exit(main())
